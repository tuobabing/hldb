# coding: utf-8

import numpy  as np
import pandas as pd

import cx_Oracle
import pymysql
import pyodbc
from sqlalchemy import create_engine

import platform
import configparser
from openpyxl import load_workbook

import os
import time
import codecs
import getpass
import zipfile
import tarfile
import ftplib as FTP

"""
A class to help with massive data operator
Support read data from oracle/mysql/hadoop to Dataframe
    and export with csv/excel with encoding=utf_8_sig and index=False
Support write dataframe to mysql table
Support ftp operator
Support zip/tar operator

自定义简单大数据操作方法的集合，
可以读取oracle/mysql/hadoop三型数据库,和将dataframe文件写入mysql表
支持ftp操作和压缩操作
"""

class hldb:
    CN_MYCONFIGFILENAME='hldb.ini'
    CN_WINDOWS="Windows"
    CN_LINUX="Linux"

    def __init__(self, inifile_dir=''):
        """
        inifile_dir 指定配置文件存放目录
        默认配置文件为当前目录下的hldb.ini
        可以向上追溯二级目录
        """
        def __getFilePath__(__rootDir):
            return os.path.join(__rootDir,self.CN_MYCONFIGFILENAME)
        def __readConfigFile__(myconfigFilePath):
            self.myconfig.read(myconfigFilePath, encoding='utf-8')
            print("读取配置文件: "+myconfigFilePath)

        self.my_platform=platform.system()
        if inifile_dir=='':
            __rootDir = os.getcwd()
        else:
            __rootDir=inifile_dir
        self.myconfig = configparser.ConfigParser()
        self.myconfigFilePath = __getFilePath__(__rootDir)

        #在当前目录查找配置文件
        if not os.path.exists(self.myconfigFilePath):
            #向上一级查找配置文件
            __rootDir=os.path.dirname(__rootDir)
            self.myconfigFilePath = __getFilePath__(__rootDir)
            if not os.path.exists(self.myconfigFilePath):
                #向上两级查找配置文件
                __rootDir=os.path.dirname(__rootDir)
                self.myconfigFilePath = __getFilePath__(__rootDir)
                #如果向上两级都未找到配置文件
                if not os.path.exists(self.myconfigFilePath):
                    print("未找到配置文件: "+self.CN_MYCONFIGFILENAME+",查找范围为'.或..或...',请执行'gen_ini()'生成配置文件")
                    return
                else:
                    __readConfigFile__(self.myconfigFilePath)
            else:
                __readConfigFile__(self.myconfigFilePath)
        else:
            __readConfigFile__(self.myconfigFilePath)

    def load_sql(self, path, args='', needCheck=False):
        """
        use placeholders to load sql from sql file
        use param needCheck=True to show total sql
        """
        if (path[-4:] != '.sql'):
            print('sql文件必须以.sql结尾')
            return ''
        __rootDir = os.getcwd()
        sqlFilePath = os.path.join(__rootDir, path)
        f = codecs.open(sqlFilePath, 'r', encoding='utf-8')
        sqls = f.readlines()
        sql = ''
        if (len(sqls) == 0):
            print('sql文件为空')
            return ''
        else:
            for line in sqls:
                sql += line
        if (args != ''):
            sql = sql.format(**args)

        simiIndex = sql.find(';')
        if ((simiIndex > -1) and (simiIndex < len(sql) - 1) and (sql.find(';', simiIndex + 1) > -1)):
            print('存在两个以上的结束符;')
            return ''

        sqlUpper = sql.upper()
        msg = ''
        if (sqlUpper.find('CREATE ')) > -1:
            msg = 'CREATE'
        elif (sqlUpper.find('COMMIT')) > -1:
            msg = 'COMMIT'
        if (msg != ''):
            msg = 'sql中含有 ' + msg + ' ，不能继续执行:'
            print(msg)
            return ''
        sql = sql.replace(';','')

        if needCheck:
            print(sql)

        return sql
    def load_sqls(self, path, args='', needCheck=False):
        """
        use placeholders list to load sqls list from one sql file
        use param needCheck=True to show total sql list
        """
        if (args == ''):
            return load_sql(path=path, needCheck=needCheck)

        if (type(args) != list):
            print('args参数必须是列表')
            return ''

        sqls = []
        for argsvalue in args:
            if (type(argsvalue) != dict):
                print('args参数内的值类型必须是字典')
                return ''

        for argsvalue in args:
            sql = self.load_sql(path=path, args=argsvalue, needCheck=False)
            sqls.append(sql)

        if (needCheck):
            for sql in sqls:
                print(sql)

        return sqls
    def write_connect(self, needPassword=False, needUser=False, writeToMysql=False):
        if writeToMysql:
            if (needUser==True):
                mysql_user = input('mysql_user:')
            if (needUser==False):
                mysql_user = self.myconfig.get('conf', 'mysql_user')
            if (needPassword==True):
                mysql_password = getpass.getpass('mysql_password:')
            if (needPassword==False):
                mysql_password = self.myconfig.get('conf', 'mysql_password')
            mysql_host = self.myconfig.get('conf', 'mysql_host')
            mysql_port = str(int(self.myconfig.get('conf', 'mysql_port')))
            mysql_schema = self.myconfig.get('conf', 'mysql_schema')

            self.__write_conn = create_engine('mysql+mysqldb://'+mysql_user+':'+mysql_password+'@'+mysql_host+':'+mysql_port+'/'+mysql_schema+'?charset=utf8')
    def connect(self, needPassword=False, needUser=False ,connToBigDB=False ,connToMysql=False):
        """
        default read user/password from ini
        modify param let read user/password from interactive(needPassword=True, needUser=True)

        default param       connect To ORACLE
        connToMysql=TRue    connect To MYSQL
        connToBigDB=TRue    connect To HADOOP
        """

        #首先判断连什么库
        if (connToMysql==False)&(connToBigDB==False):

            #然后判断读从哪里读配置
            if (needUser==True):
                oracle_user = input('oracle_user:')
            if (needUser==False):
                oracle_user = self.myconfig.get('conf', 'oracle_user')
            if (needPassword==True):
                oracle_password = getpass.getpass('oracle_password:')
            if (needPassword==False):
                oracle_password = self.myconfig.get('conf', 'oracle_password')

            #然后从配置读取基础配置信息
            oracle_dbip = self.myconfig.get('conf', 'oracle_dbip')
            oracle_dbname = self.myconfig.get('conf', 'oracle_dbname')
            db = oracle_user + '/' + oracle_password + '@' + oracle_dbip + ':1521/' + oracle_dbname

            #最后建立连接
            self.__conn = cx_Oracle.connect(db + '', encoding = 'UTF-8', nencoding='UTF-8')
            self.__cursor = self.__conn.cursor()
            print('oracle connected')

        elif (connToMysql==True)&(connToBigDB==False):
            if (needUser==True):
                mysql_user = input('mysql_user:')
            if (needUser==False):
                mysql_user = self.myconfig.get('conf', 'mysql_user')
            if (needPassword==True):
                mysql_password = getpass.getpass('mysql_password:')
            if (needPassword==False):
                mysql_password = self.myconfig.get('conf', 'mysql_password')

            mysql_host = self.myconfig.get('conf', 'mysql_host')
            mysql_port = int(self.myconfig.get('conf', 'mysql_port'))
            mysql_schema = self.myconfig.get('conf', 'mysql_schema')

            self.__conn = pymysql.connect(host=mysql_host,port=mysql_port,database=mysql_schema,charset='utf8',user=mysql_user,password=mysql_password)
            self.__cursor = self.__conn.cursor()
            print('mysql connected')

        elif (connToMysql==False)&(connToBigDB==True):
            if (needUser==True):
                impala_user = input('impala_user:')
            if (needUser==False):
                impala_user = self.myconfig.get('conf', 'impala_user')
            if (needPassword==True):
                impala_password = getpass.getpass('impala_password:')
            if (needPassword==False):
                impala_password = self.myconfig.get('conf', 'impala_password')

            impala_host = self.myconfig.get('conf', 'impala_host')
            impala_port = int(self.myconfig.get('conf', 'impala_port'))
            impala_driver_name = self.myconfig.get('conf', 'impala_driver_name')

            self.__conn = self.__connImpala(ip=impala_host, port=impala_port, user=impala_user, password=impala_password,impala_driver_name=impala_driver_name)
            self.__cursor = self.__conn.cursor()
            self.__cursor.execute('show databases')
            print('hadoop connected')

        elif (connToMysql==True)&(connToBigDB==True):
            print('一个实例不能同时连接多个数据库,请检查输入条件')

        else:
            print('异常条件')

    def __connImpala(self, ip=None, port=None, user=None, password=None,impala_driver_name=None):
        if self.my_platform==self.CN_WINDOWS:
            driverParas = 'DRIVER={};Host={};Port={};AuthMech=3;UID={};PWD={}'.format(impala_driver_name, ip, port, user,password)
        elif self.my_platform==self.CN_LINUX:
            driverParas = 'DSN={};Host={};Port={};AuthMech=3;UID={};PWD={}'.format(impala_driver_name, ip, port, user,password)
        else:
            print("目前只兼容Windows和linux环境")
        conn = pyodbc.connect(driverParas, autocommit=True)
        return conn

    def read_sql(self, sql, fileprefix='', filesuffix='', sheetname=''):
        """
        read data from db with sql,return dataframe
        fileprefix is datafile's name before '.'
        filesuffix is 'csv' or 'xlsx',not 'xls'
        default data filesuffix is 'csv'
        sheetname only used when filesuffix='xlsx'
        """
        return self.__read_sql(sql,fileprefix=fileprefix, filesuffix=filesuffix, sheetname=sheetname)
    def write_db(self, src_df, table_name, mysql_schema, db='mysql', if_exists='append', chunksize=1000000):
        """
        write data from df to db,only mysql now
        """
        if db=='mysql':
            pd.io.sql.to_sql(frame=src_df,name=table_name,con=self.__write_conn,
                         schema=mysql_schema, if_exists=if_exists,index=False,chunksize=chunksize)
            self.__write_conn.dispose()
            print("写库完成，主动断开写库session")
        if db=='oracle':
            print("写库完成，主动断开写库session")

    def __read_sql(self, sql, fileprefix='', filesuffix='', sheetname='', batchsuffix=''):
        sql = sql.replace(';','')
        starttime = time.time()
        arr = pd.read_sql(sql, self.__conn)
        print('sql executed time : ' + format(time.time() - starttime, '0.3f') + 's')
        df = pd.DataFrame(arr)
        timestamp = time.strftime('%Y-%m-%d %H-%M-%S',time.localtime())
        filename=''
        if (fileprefix != ''):
            filename = fileprefix
        #batchsuffix:read_sqls批量生成文件的时候用以区分生成的文件名
        filename += batchsuffix
        if (not (fileprefix == '' and filesuffix == '' and sheetname == '' and batchsuffix == '')):
            if (filesuffix == 'xlsx' and fileprefix != ''):
                if (sheetname == ''):
                    sheetname = 'sheet1'
                filename += '.xlsx'
                if (os.path.exists(filename)):
                    book = load_workbook(filename)
                    sheets = book.sheetnames
                    if (sheetname in sheets):
                        idx = sheets.index(sheetname)
                        book.remove(book.worksheets[idx])
                    writer = pd.ExcelWriter(filename, engine='openpyxl')
                    writer.book = book
                    df.to_excel(writer, sheetname)
                    writer.save()
                else:
                    writer = pd.ExcelWriter(filename)
                    df.to_excel(writer, sheetname)
                    writer.save()
                self.__modify_columns_width(df, filename, sheetname)
            else :
                df.to_csv(filename + '.csv',index=False, encoding='utf_8_sig')
        return df
    def __modify_columns_width(self, df, filename, sheetname):
        wb = load_workbook(filename)
        ws_list = wb.sheetnames
        for i in range(len(ws_list)):
            ws = wb[ws_list[i]]
            if (ws_list[i] != sheetname):
                continue
            df_len = df.apply(lambda x:[(len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i)) for i in x], axis = 0)
            df_len_max = df_len.apply(lambda x:max(x), axis=0)
            j = list(df.columns)
            j.insert(0, '')
            for i in df.columns:
                column_letter = [chr(j.index(i) + 65) if j.index(i) <= 25 else 'A'+chr(j.index(i) - 26 + 65)][0]
                columns_length = (len(str(i).encode('utf-8')) - len(str(i))) / 2 + len(str(i))
                data_max_length = df_len_max[i]
                column_width = [data_max_length if columns_length < data_max_length else columns_length][0]
                column_width = [column_width if column_width > 15 else 15][0]
                ws.column_dimensions['{}'.format(column_letter)].width = column_width
        wb.save(filename=filename)
    def read_sqls(self, sqls, fileprefix='', filesuffix='', sheetname=''):
        dfs = []
        batchsuffix = ''
        isFileCreated = not (fileprefix == '' and filesuffix == '' and sheetname == '')
        for i in range(len(sqls)):
            if (isFileCreated):
                batchsuffix = '_'+str(1+i)
            dfs.append(self.__read_sql(sqls[i],fileprefix=fileprefix,filesuffix=filesuffix,sheetname=sheetname,batchsuffix=batchsuffix))
        return dfs
    def db_write(self, needPassword=True, needUser=True ,connToBigDB=False ,connToMysql=False):
        if connToMysql:
            con4write = create_engine('mysql+mysqldb://test:testpassword2020@localhost:3306/baostock?charset=utf8')
        pass
    def ftp_update(self,remotefilepath,localfilename, needUser=False, needPassword=False):
        """
        used for update local file to ftp's remotefilepath

        default read user/password from ini
        modify param let read user/password from interactive(needPassword=True, needUser=True)
        """
        if (needUser):
            ftp_username = input('ftp_user:')
        else:
            ftp_username = self.myconfig.get('conf', 'ftp_user')

        if (needPassword):
            ftp_password = getpass.getpass('ftp_password:')
        else:
            ftp_password = self.myconfig.get('conf', 'ftp_password')
        ftp_ip = self.myconfig.get('conf', 'ftp_ip')
        ftp_port = self.myconfig.get('conf', 'ftp_port')
        bufsize = self.myconfig.get('conf', 'ftp_bufsize')

        ftp=FTP.FTP()
        ftp.encoding = self.myconfig.get('conf', 'ftp_encoding')
        ftp.connect(ftp_ip,ftp_port)

        ftp.login(ftp_username,ftp_password)
        ftp.cwd(remotefilepath)
        file=open(localfilename,'rb')
        ftp.storbinary('STOR '+localfilename,file,bufsize)
        file.close()

        print("上传后FTP目录下文件为: ")
        ftp.dir()
        ftp.quit()
    def ftp_download(self,remotefilepath,remotefilename,localfilename, needUser=False, needPassword=False):
        """
        used for download remote file on ftp to local file

        default read user/password from ini
        modify param let read user/password from interactive(needPassword=True, needUser=True)
        """
        ftp_port = self.myconfig.get('conf', 'ftp_port')
        bufsize = self.myconfig.get('conf', 'ftp_bufsize')
        if (needUser):
            ftp_username = input('ftp_user:')
        else:
            ftp_username = self.myconfig.get('conf', 'ftp_user')
        if (needPassword):
            ftp_password = getpass.getpass('ftp_password:')
        else:
            ftp_password = self.myconfig.get('conf', 'ftp_password')

        ftp_ip = self.myconfig.get('conf', 'ftp_ip')
        ftp=FTP.FTP()
        ftp.encoding = self.myconfig.get('conf', 'ftp_encoding')
        ftp.connect(ftp_ip,ftp_port)

        ftp.login(ftp_username,ftp_password)
        ftp.cwd(remotefilepath)
        file=open(localfilename,'wb')
        ftp.retrbinary('RETR '+remotefilename,file.write,bufsize)
        file.close()
        ftp.quit()

        print("下载后当前目录文件：")
        print(os.listdir())

    def close(self):
        """
        Don't forget to use close() at last!
        use it to close connect to db
        and clear any plainetxt password in hldb.ini
        """
        self.__conn.close()
        print('hldb disconnected')
        self.myconfig.set('conf', 'oracle_password', '')
        self.myconfig.set('conf', 'ftp_password', '')
        self.myconfig.set('conf', 'impala_password', '')
        self.myconfig.set('conf', 'mysql_password', '')
        self.myconfig.write(open(self.myconfigFilePath, 'w'))
        print('password is clear')
    def gen_ini(self):
        """
        Generate a profile in the current directory for next
        please modify it for your env
        """
        Isgen=True
        if os.path.exists(self.CN_MYCONFIGFILENAME):
            print(self.CN_MYCONFIGFILENAME+"已经存在，是否需要覆盖? y/n")
            if (str.upper(input()) != 'Y'):
                Isgen=False
        if Isgen==True:
            config = configparser.ConfigParser()
            config['conf'] = {
                            'oracle_dbip' : 'x.x.x.x',
                            'oracle_dbname' : '',
                            'oracle_user' : '',
                            'oracle_password' : '',
                            'ftp_username' : '',
                            'ftp_password' : '',
                            'ftp_ip' : 'x.x.x.x',
                            'ftp_port' : '21',
                            'ftp_bufsize' : '1024',
                            'impala_host' : 'x.x.x.x',
                            'impala_port' : '21050',
                            'impala_user' : '',
                            'impala_password' : '',
                            'impala_driver_name' : 'Cloudera ODBC Driver for Impala',
                            'mysql_host' : 'x.x.x.x',
                            'mysql_port' : '3306',
                            'mysql_schema' : '',
                            'mysql_user' : '',
                            'mysql_password' : ''
                             }
            with open(self.CN_MYCONFIGFILENAME, 'w') as configfile:
                config.write(configfile)
            print("已在当前目录下生成配置文件"+self.CN_MYCONFIGFILENAME+",请修改该文件的配置以方便使用")
    def compress_file(self,src_name,compress_name,mode='zip'):
        """
        compress a file to 'zip' or 'tar' on param
        """
        if ((src_name=='') or (compress_name=='')):
            print("请输入待压缩或压缩后的文件名或目录名")
            return
        if  os.path.exists(src_name) != True:
            print("待压缩文件不存在")
            return
        if mode=='zip':
            compress_name+='.zip'
            if os.path.isfile(src_name):
                with zipfile.ZipFile(compress_name, 'w') as z:
                    z.write(src_name)
            else:
                with zipfile.ZipFile(compress_name, 'w') as z:
                    for root, dirs, files in os.walk(src_name):
                        for single_file in files:
                            if single_file != compress_name:
                                filepath = os.path.join(root, single_file)
                                z.write(filepath)
                z.close()
        elif mode=='tar':
            compress_name+='.tar'
            if os.path.isfile(src_name):
                with tarfile.open(compress_name, 'w') as tar:
                    tar.add(src_name)
            else:
                with tarfile.open(compress_name, 'w') as tar:
                    for root, dirs, files in os.walk(src_name):
                        for single_file in files:
                            if single_file != compress_name:
                                filepath = os.path.join(root, single_file)
                                tar.add(filepath)
                z.close()
        else:
            print("目前仅支持zip/tar压缩方式")

        return compress_name
    def check_df_null(self,check_df):
        return check_df[check_df.isnull().T.any()]
    def help(self, parameter='', help_list=False):
        parameter=parameter.strip()
        if len(parameter)==0:
            help_list=True
        else:
            help_list=False
            print("How to use " + "," + parameter + " :")
            print("-----------")
        #####################################
        #列表区
        #####################################
        if help_list:
            print('''
    常用命令检索：help('hldb')
    -----------
    hldb              : hldb包的使用方法
    joblib            : 使用并行joblib命令,兼容Windows
    df_parallel_apply : 使用并行apply()命令
    df_iterrows       : DataFrame循环使用命令
    df_to_csv         : 导出文件不显示索引并正常显示中文命令
    df_from_ndarry    : ndarry转DataFrame命令
    df_from_list      : list转dataframe命令
    df_todatetime     : 日期字符类型转为日期类型(画图可显示)
    is_null           : 判断空的几种方法
    percentile        : 计算分位数值和获取某个值在某个系列的分位数
    plt_rotation      : X轴坐标旋转命令
    plt_annotate      : 图点加标记命令
    plt_3d            : 3D绘画命令
    plt_chinese       : 画图显示中文命令
    plt_to_big        : 画图显示大坐标命令
    impala_drive      : 显示impala安装驱动步骤
    show_mem          : 显示内存使用情况
    import_hive_table : 脚本导入hive数据的关键步骤
    ''')
        #####################################
        #字典区
        #####################################
        if parameter=='df_parallel_apply':
            print('''
    from pandarallel import pandarallel
    pandarallel.initialize()
    df.parallel_apply(func,**kwds,axis=1)
    df[["aa","bb"]]=df.parallel_apply(func,axis=1,result_type="expand")''')
        #####################################
        if parameter=='joblib':
            print('''
    from math import sqrt
    from joblib import Parallel, delayed
    import time
    def mysqrt(i):
        time.sleep(2)
        return sqrt(i**2)
    t_start=time.time()
    #Parallel(n_jobs=4, verbose=20)(delayed(sqrt)(i**2) for i in range(10))
    Parallel(n_jobs=4)(delayed(mysqrt)(i**2) for i in range(10))
    tmptime2=time.time()-t_start
    print(tmptime2)''')
        #####################################
        if parameter=='df_todatetime':
            print('''
    df["TRADINGDAY"]=pd.to_datetime(df.TRADINGDAY,format="%Y%m%d")''')
        #####################################
        if parameter=='df_iterrows':
            print('''
    for index,row in df.iterrows()''')
        #####################################
        if parameter=='df_from_ndarry':
            print('''
    df=pd.DataFrame(ndarry.reshape(-1,4),columns=["xx","yy"])''')
        #####################################
        if parameter=='df_to_csv':
            print('''
    df.to_csv("df.csv",index=False,encoding='utf_8_sig')''')
        #####################################
        if parameter=='plt_rotation':
            print('''
    for xtick in ax.get_xticklabels():
        xtick.set_rotation(50)''')
        #####################################
        if parameter=='plt_3d':
            print('''
    import matplotlib.pyplot as plt
    from mpl_toolkits.mplot3d import Axes3D
    fig=plt.figure(figsize=(9,9))
    ax=fig.add_subplot(111,projection='3d')
    ax.scatter(result["x"],result["y"],result["z"])''')
        #####################################
        if parameter=='show_mem':
            print('''
    import psutil
    psutil.virtual_memory().percent
    #shell:while true; do free; sleep 10; done
    ''')
        #####################################
        if parameter=='plt_annotate':
            print('''
    for index,row in df.iterrows():
        plt.annotate(xy=(index,row.PRICE),s=row.VOLUME)''')
        #####################################
        if parameter=='percentile':
            print('''
    from scipy import stats
    stats.percentileofscore(a, score, kind='rank')
    import numpy as np
    np.percentile()''')
        #####################################
        if parameter=='impala_drive':
            print('''
    在Windows上安装ClouderaImpalaODBC64-2.5.41.msi
    在Linux上安装ClouderaImpalaODBC-2.5.41.1029-1.el7.x86_64.rpm或更高版本
    Windows不需要配置文件,Linux配置文件类似.odbc.ini
    [Cloudera ODBC Driver for Impala]
    Driver=/opt/cloudera/impalaodbc/lib/64/libclouderaimpalaodbc64.so
    HOST=x.x.x.x
    PORT=21050
    AuthMech=3''')
        #####################################
        if parameter=='is_null':
            print('''
    方法1：字符串判断为空:
    x=None
    if x is None:
        print("x is None")

    方法2：np判断为空:
    >>> np.isnan(np.nan)
    True
    >>> np.isnan(np.inf)
    False
    >>> np.isnan([np.log(-1.),1.,np.log(0)])
    array([ True, False, False])

    方法3：Series判断为空:
    y=pd.Series([])
    if y.empty:
        print("为空")
        ''')
        #####################################
        if parameter=='hldb':
            print('''
    #导入包
    import hldb as hldb

    #生成配置文件模板
    gen_ini()

    #初始化,指定配置文件位置，否则在最近两级上级目录查找
    hl = hldb.hldb(inifile_dir='D:\jupyterNotebook')
    hl = hldb.hldb(inifile_dir='/home/trade/jupyterHome')

    #连接数据库
    hl.connect(needPassword=True,needUser=True)
    hd.connect(needPassword=True,needUser=True,connToBigDB=True)

    #sql配置参数
    args = {'startday':'20181001', 'endday':'20181101','clienttype':'= 0'}

    #读取完整sql
    sql = hl.load_sql(path='test.sql',args=args, needCheck=False)

    #执行sql，并写入本地文件
    df = hl.read_sql(sql=sql,fileprefix='', filesuffix='', sheetname='')

    #df落地csv文件
    #df.to_csv("df.csv",,index=False, encoding='utf_8_sig')

    #关闭连接，清除配置文件中的密码记录
    hl.close()

    # ftp上传
    ftp_update(remotefilepath,localfilename, needUser=False, needPassword=False)
    # ftp下载
    ftp_download(remotefilepath,remotefilename,localfilename, needUser=False, needPassword=False)
    # 文件压缩
    compress_file(src_name,compress_name,mode='zip')
        ''')
        #####################################
        if parameter=='plt_chinese':
            print('''
    import matplotlib.pyplot as plt
    plt.rcParams["font.sans-serif"]=["Microsoft YaHei"]
        ''')
        #####################################
        if parameter=='plt_to_big':
            print('''
    import matplotlib.pyplot as plt
    import matplotlib.ticker as mtick
    #设置文字
    ax.set_title("xx",fontsize=28)
    ax.set_xlabel("xx",fontsize=28)
    ax.set_ylabel("xx",fontsize=28)

    #标签变大，可能导致显示不全
    ax.legend(prop={'size':22})
    #坐标刻度变大,但不能将图右侧tick变大
    ax.tick_params(labelsize=28)
    #或者
    plt.yticks(fontsize=16)

    #y轴显示百分比
    fmt='%.2f%%'
    yticks=mtick.FormatStrFormatter(fmt)
    ax.yaxis.set_major_formatter(yticks)
        ''')
        #####################################
        if parameter=='df_from_list':
            print('''
    columnsX=["x","y","a","b","c"]
    df=pd.DataFrame(columns=columnsX)
    tmplist=[]
    tmplist.append([1,2,3,4,5])
    df=df.append(pd.DataFrame(tmplist,columns=columnsX))''')
        #####################################
        if parameter=='import_hive_table':
            print('''
    dos2unix $csv_full_name;
    scp $csv_full_name impala@remmote:/tmp/hadoop_table/
    ssh impala@remmote "hostname; source /var/lib/impala/.bash_profile;
                        hdfs dfs -mkdir -p /tmp/hsperfdata_mportal/$table_name;
                        hdfs dfs -ls /tmp/hsperfdata_mportal;
                        hdfs dfs -put /tmp/hadoop_table/$csv_name /tmp/hsperfdata_mportal/$table_name/"
    ''')
        #####################################
